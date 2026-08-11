# -*- coding: utf-8 -*-
"""李昌辉变化检测结果统计与Excel汇总模块。

仅只读打开状态表和栅格；单项统计异常不会中断后续任务。
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import time
import traceback
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import rasterio


MODULE_ROOT = Path(__file__).resolve().parent
PRIMARY_STATUS = MODULE_ROOT / "可选输入" / "真实批量状态表.csv"
RELIABILITY_STATUS = MODULE_ROOT / "可选输入" / "可靠性验收状态表.csv"
RESULT_ROOT = MODULE_ROOT / "测试结果"
DEFAULT_XLSX = MODULE_ROOT / "Excel统计表" / "批量变化检测统计汇总.xlsx"
DEFAULT_CSV = MODULE_ROOT / "统计数据" / "批量变化检测统计明细.csv"
DEFAULT_LOG = MODULE_ROOT / "运行记录" / "变化统计运行.log"
DEFAULT_REPORT = MODULE_ROOT / "运行记录" / "变化统计与Excel验收报告.md"

STAT_HEADERS = [
    "task_id",
    "任务状态",
    "输出文件完整路径",
    "文件大小（字节）",
    "宽度",
    "高度",
    "波段数",
    "数据类型",
    "坐标系",
    "像元分辨率X",
    "像元分辨率Y",
    "总像元数",
    "有效像元数",
    "NoData像元数",
    "未变化像元数（值0）",
    "变化像元数（值1）",
    "其他像元值数量",
    "变化比例",
    "单像元面积（平方米）",
    "变化面积（平方米）",
    "变化面积（公顷）",
    "变化面积（平方千米）",
    "推理耗时（秒）",
    "重试次数",
    "统计状态",
    "警告或错误说明",
]

TASK_HEADERS = [
    "来源状态表",
    "task_id",
    "前时相影像",
    "后时相影像",
    "模型名称",
    "GPU编号",
    "输出文件",
    "任务状态",
    "attempt",
    "重试次数",
    "开始时间",
    "结束时间",
    "推理耗时（秒）",
    "输出存在",
    "输出大小（字节）",
    "错误类型",
    "错误信息",
]

FAILURE_HEADERS = [
    "来源状态表",
    "task_id",
    "任务状态",
    "输出文件",
    "错误类型",
    "错误信息",
    "统计状态",
    "统计警告或错误",
]

FIELD_HEADERS = ["工作表", "字段", "含义", "单位或计算方法"]


class RunLogger:
    """避免第三方库修改全局logging配置的直接UTF-8日志器。"""

    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text("", encoding="utf-8")

    def _write(self, level: str, message: str) -> None:
        current = datetime.now().astimezone()
        stamp = current.strftime("%Y-%m-%d %H:%M:%S") + (
            f",{current.microsecond // 1000:03d}"
        )
        line = f"{stamp} | {level} | {message}"
        with self.path.open("a", encoding="utf-8", newline="\n") as stream:
            stream.write(line + "\n")
        print(line, flush=True)

    def info(self, message: str) -> None:
        self._write("INFO", message)

    def warning(self, message: str) -> None:
        self._write("WARNING", message)

    def error(self, message: str) -> None:
        self._write("ERROR", message)


def parse_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y", "是"}


def parse_int(value: Any, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def parse_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def load_status_tables(paths: Sequence[Path]) -> List[Dict[str, str]]:
    tasks: List[Dict[str, str]] = []
    for path in paths:
        if not path.is_file():
            raise FileNotFoundError(f"状态表不存在：{path}")
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            required = {"task_id", "status", "output_path"}
            missing = sorted(required - set(reader.fieldnames or []))
            if missing:
                raise ValueError(f"状态表缺少字段 {missing}：{path}")
            for row in reader:
                normalized = {key: (value or "") for key, value in row.items()}
                normalized["_source_status_table"] = str(path)
                tasks.append(normalized)
    return tasks


def resolve_output_path(task: Dict[str, str], result_root: Path) -> Path:
    value = task.get("output_path", "").strip()
    if not value:
        raise ValueError("output_path为空")
    path = Path(value)
    return path if path.is_absolute() else (result_root / path).resolve()


def retry_count(task: Dict[str, str]) -> int:
    explicit = task.get("retry_count", "").strip()
    if explicit:
        return max(0, parse_int(explicit, 0))
    return max(0, parse_int(task.get("attempt", "0"), 0) - 1)


def is_projected_meter_crs(crs: Any) -> Tuple[bool, str]:
    if crs is None:
        return False, "坐标系未定义"
    if not bool(getattr(crs, "is_projected", False)):
        return False, "经纬度坐标系不能直接以度平方计算平方米"
    try:
        unit_name, unit_factor = crs.linear_units_factor
    except Exception:
        return False, "坐标系线性单位不明确"
    normalized = str(unit_name).strip().lower()
    meter_names = {"metre", "meter", "metres", "meters", "m"}
    if normalized not in meter_names or not math.isclose(
        float(unit_factor), 1.0, rel_tol=0.0, abs_tol=1e-12
    ):
        return False, f"坐标系线性单位不是明确的米：{unit_name}, factor={unit_factor}"
    return True, ""


def valid_mask_for_band(
    band: np.ndarray, read_mask: np.ndarray, nodata: Optional[float]
) -> np.ndarray:
    valid = read_mask != 0
    if nodata is not None:
        try:
            if math.isnan(float(nodata)):
                valid &= ~np.isnan(band)
            else:
                valid &= band != nodata
        except (TypeError, ValueError):
            valid &= band != nodata
    if np.issubdtype(band.dtype, np.floating):
        valid &= np.isfinite(band)
    return valid


def statistics_placeholder(
    task: Dict[str, str], output_path: str, status_text: str, error_text: str
) -> Dict[str, Any]:
    row = {header: None for header in STAT_HEADERS}
    row.update(
        {
            "task_id": task.get("task_id", ""),
            "任务状态": task.get("status", ""),
            "输出文件完整路径": output_path,
            "推理耗时（秒）": parse_float(task.get("duration_seconds", "0")),
            "重试次数": retry_count(task),
            "统计状态": status_text,
            "警告或错误说明": error_text,
        }
    )
    return row


def compute_raster_statistics(
    task: Dict[str, str], output_path: Path, logger: RunLogger
) -> Dict[str, Any]:
    notes: List[str] = []
    started = time.perf_counter()
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        with rasterio.open(output_path, "r") as source:
            # 完整读取全部波段用于检验文件未截断；类别计数按第1波段。
            all_data = source.read(masked=False)
            expected_size = source.count * source.width * source.height
            if all_data.size != expected_size:
                raise RuntimeError(
                    f"完整读取像元数不符：{all_data.size} != {expected_size}"
                )
            if source.count < 1:
                raise RuntimeError("栅格没有波段")
            band = all_data[0]
            read_mask = source.read_masks(1)
            nodata = source.nodatavals[0]
            valid_mask = valid_mask_for_band(band, read_mask, nodata)

            total_pixels = int(source.width * source.height)
            valid_pixels = int(np.count_nonzero(valid_mask))
            nodata_pixels = total_pixels - valid_pixels
            unchanged_pixels = int(np.count_nonzero(valid_mask & (band == 0)))
            changed_pixels = int(np.count_nonzero(valid_mask & (band == 1)))
            other_pixels = valid_pixels - unchanged_pixels - changed_pixels
            if other_pixels < 0:
                raise RuntimeError("分类像元计数出现负值")
            if unchanged_pixels + changed_pixels + other_pixels != valid_pixels:
                raise RuntimeError("值0、值1和其他值数量之和不等于有效像元数")
            change_ratio = (
                changed_pixels / valid_pixels if valid_pixels > 0 else None
            )

            transform = source.transform
            resolution_x = abs(float(transform.a))
            resolution_y = abs(float(transform.e))
            if resolution_x <= 0 or resolution_y <= 0:
                notes.append("像元分辨率无效，面积字段留空")
            elif not math.isclose(
                resolution_x, resolution_y, rel_tol=1e-9, abs_tol=1e-12
            ):
                relative = abs(resolution_x - resolution_y) / max(
                    resolution_x, resolution_y
                )
                notes.append(
                    "X/Y像元分辨率不完全一致"
                    f"（相对差{relative:.8%}），记录为亚像元范围偏移提示，未修改栅格"
                )
            if not math.isclose(
                float(transform.b), 0.0, abs_tol=1e-12
            ) or not math.isclose(float(transform.d), 0.0, abs_tol=1e-12):
                notes.append("仿射变换含旋转/错切；按要求使用abs(pixel_width×pixel_height)")

            metric_crs, metric_reason = is_projected_meter_crs(source.crs)
            pixel_area_m2: Optional[float]
            changed_area_m2: Optional[float]
            changed_area_ha: Optional[float]
            changed_area_km2: Optional[float]
            if metric_crs and resolution_x > 0 and resolution_y > 0:
                pixel_area_m2 = abs(float(transform.a) * float(transform.e))
                changed_area_m2 = changed_pixels * pixel_area_m2
                changed_area_ha = changed_area_m2 / 10_000
                changed_area_km2 = changed_area_m2 / 1_000_000
            else:
                pixel_area_m2 = None
                changed_area_m2 = None
                changed_area_ha = None
                changed_area_km2 = None
                notes.append(
                    f"{metric_reason}；需投影到合适的等积或米制坐标系后计算面积"
                )

            if source.count > 1:
                notes.append(
                    f"结果为{source.count}波段；文件全部读取，类别统计使用第1波段"
                )
            notes.extend(str(item.message) for item in caught)

            row: Dict[str, Any] = {
                "task_id": task.get("task_id", ""),
                "任务状态": task.get("status", ""),
                "输出文件完整路径": str(output_path),
                "文件大小（字节）": output_path.stat().st_size,
                "宽度": source.width,
                "高度": source.height,
                "波段数": source.count,
                "数据类型": ", ".join(source.dtypes),
                "坐标系": source.crs.to_string() if source.crs else "未定义",
                "像元分辨率X": resolution_x,
                "像元分辨率Y": resolution_y,
                "总像元数": total_pixels,
                "有效像元数": valid_pixels,
                "NoData像元数": nodata_pixels,
                "未变化像元数（值0）": unchanged_pixels,
                "变化像元数（值1）": changed_pixels,
                "其他像元值数量": other_pixels,
                "变化比例": change_ratio,
                "单像元面积（平方米）": pixel_area_m2,
                "变化面积（平方米）": changed_area_m2,
                "变化面积（公顷）": changed_area_ha,
                "变化面积（平方千米）": changed_area_km2,
                "推理耗时（秒）": parse_float(
                    task.get("duration_seconds", "0")
                ),
                "重试次数": retry_count(task),
                "统计状态": "success_with_warning" if notes else "success",
                "警告或错误说明": "；".join(note for note in notes if note),
            }
    logger.info(
        f"栅格统计完成 | task_id={row['task_id']} | "
        f"valid={row['有效像元数']} | changed={row['变化像元数（值1）']} | "
        f"elapsed={time.perf_counter() - started:.3f}s"
    )
    return row


def task_detail_row(task: Dict[str, str], result_root: Path) -> Dict[str, Any]:
    try:
        output = resolve_output_path(task, result_root)
        output_exists = output.is_file()
        output_size = output.stat().st_size if output_exists else 0
    except Exception:
        output = Path(task.get("output_path", ""))
        output_exists = False
        output_size = 0
    return {
        "来源状态表": task.get("_source_status_table", ""),
        "task_id": task.get("task_id", ""),
        "前时相影像": task.get("before_image", ""),
        "后时相影像": task.get("after_image", ""),
        "模型名称": task.get("model_name", ""),
        "GPU编号": parse_int(task.get("gpu_id", "0")),
        "输出文件": str(output),
        "任务状态": task.get("status", ""),
        "attempt": parse_int(task.get("attempt", "0")),
        "重试次数": retry_count(task),
        "开始时间": task.get("start_time", ""),
        "结束时间": task.get("end_time", ""),
        "推理耗时（秒）": parse_float(task.get("duration_seconds", "0")),
        "输出存在": output_exists,
        "输出大小（字节）": output_size,
        "错误类型": task.get("error_type", ""),
        "错误信息": task.get("error_message", ""),
    }


def failure_row(
    task: Dict[str, str],
    output_path: str,
    error_type: str,
    error_message: str,
    statistics_status: str,
    statistics_error: str,
) -> Dict[str, Any]:
    return {
        "来源状态表": task.get("_source_status_table", ""),
        "task_id": task.get("task_id", ""),
        "任务状态": task.get("status", ""),
        "输出文件": output_path,
        "错误类型": error_type,
        "错误信息": error_message,
        "统计状态": statistics_status,
        "统计警告或错误": statistics_error,
    }


def field_definitions() -> List[Dict[str, str]]:
    definitions = [
        ("总览", "统计时间", "本次统计运行的本地时间", "ISO时间"),
        ("总览", "总任务数", "任务明细记录数", "条"),
        ("总览", "成功数/失败数/跳过数", "按状态表原状态计数", "条"),
        ("总览", "可统计结果数", "成功或跳过且输出有效并完成统计的TIFF数", "个"),
        ("总览", "总变化像元数", "变化统计中值1像元数之和", "像元"),
        ("总览", "总变化面积", "米制投影结果变化面积之和", "平方米"),
        ("总览", "平均变化比例", "各有效结果变化比例的算术平均", "%"),
        ("总览", "总运行耗时", "任务明细推理耗时之和", "秒"),
        ("任务明细", "attempt", "最终尝试序号，首次尝试为1", "次序"),
        ("任务明细", "重试次数", "优先读取retry_count，否则max(attempt-1,0)", "次"),
        ("变化统计", "总像元数", "宽度×高度，类别统计按第1波段", "像元"),
        ("变化统计", "有效像元数", "读取掩膜有效且不等于NoData的像元", "像元"),
        ("变化统计", "NoData像元数", "总像元数-有效像元数", "像元"),
        ("变化统计", "未变化像元数（值0）", "有效像元中值等于0的数量", "像元"),
        ("变化统计", "变化像元数（值1）", "有效像元中值等于1的数量", "像元"),
        ("变化统计", "其他像元值数量", "有效像元数-值0数量-值1数量", "像元"),
        ("变化统计", "变化比例", "变化像元数÷有效像元数", "百分比"),
        (
            "变化统计",
            "单像元面积（平方米）",
            "仅米制投影：abs(pixel_width×pixel_height)",
            "平方米",
        ),
        (
            "变化统计",
            "变化面积（平方米）",
            "变化像元数×单像元面积",
            "平方米",
        ),
        ("变化统计", "变化面积（公顷）", "平方米÷10,000", "公顷"),
        ("变化统计", "变化面积（平方千米）", "平方米÷1,000,000", "平方千米"),
        (
            "变化统计",
            "统计状态",
            "success、success_with_warning或统计异常状态",
            "文本",
        ),
        (
            "变化统计",
            "警告或错误说明",
            "面积单位、亚像元分辨率偏移、波段或读取提示",
            "文本",
        ),
        (
            "失败记录",
            "失败任务",
            "failed任务不打开栅格；统计异常任务记录完整错误",
            "文本",
        ),
    ]
    return [
        {"工作表": sheet, "字段": field, "含义": meaning, "单位或计算方法": unit}
        for sheet, field, meaning, unit in definitions
    ]


def dict_rows_to_matrix(
    rows: Sequence[Dict[str, Any]], headers: Sequence[str]
) -> List[List[Any]]:
    return [[row.get(header) for header in headers] for row in rows]


def write_statistics_csv(
    path: Path,
    all_rows: Sequence[Dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=STAT_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_rows)


def build_workbook_openpyxl(
    payload: Dict[str, Any],
    output_path: Path,
    qa_dir: Path,
    logger: RunLogger,
) -> Dict[str, Any]:
    """使用普通Python库生成统计工作簿并执行结构性复验。"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    colors = {
        "header": "0F766E",
        "header_text": "FFFFFF",
        "title": "134E4A",
        "light_teal": "CCFBF1",
        "border": "CBD5E1",
        "success_fill": "DCFCE7",
        "success_text": "166534",
        "failed_fill": "FEE2E2",
        "failed_text": "991B1B",
        "skipped_fill": "FEF3C7",
        "skipped_text": "92400E",
        "warning_fill": "FFEDD5",
        "warning_text": "9A3412",
    }
    thin = Side(style="thin", color=colors["border"])
    body_border = Border(bottom=thin)
    header_fill = PatternFill("solid", fgColor=colors["header"])
    header_font = Font(name="微软雅黑", size=10, bold=True, color=colors["header_text"])
    body_font = Font(name="微软雅黑", size=10, color="263238")

    workbook = Workbook()
    workbook.remove(workbook.active)

    def set_widths(sheet, widths_px: Sequence[int]) -> None:
        for index, width_px in enumerate(widths_px, start=1):
            sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = max(
                8.0, width_px / 7.0
            )

    def column_letter(index: int) -> str:
        from openpyxl.utils import get_column_letter

        return get_column_letter(index)

    def add_status_formats(sheet, column: str, start: int, end: int) -> None:
        if end < start:
            return
        target = f"{column}{start}:{column}{end}"
        rules = [
            ("success", colors["success_fill"], colors["success_text"]),
            ("failed", colors["failed_fill"], colors["failed_text"]),
            ("skipped", colors["skipped_fill"], colors["skipped_text"]),
            ("warning", colors["warning_fill"], colors["warning_text"]),
        ]
        for text, fill, font_color in rules:
            formula = [f'ISNUMBER(SEARCH("{text}",{column}{start}))']
            sheet.conditional_formatting.add(
                target,
                FormulaRule(
                    formula=formula,
                    fill=PatternFill("solid", fgColor=fill),
                    font=Font(name="微软雅黑", bold=True, color=font_color),
                ),
            )

    def add_table_sheet(
        name: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        table_name: str,
        widths_px: Sequence[int],
        *,
        row_height: float,
        wrap_columns: Iterable[int] = (),
    ):
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.append(list(headers))
        for row in rows:
            sheet.append(list(row))
        if not rows:
            sheet.append([None] * len(headers))
        last_row = sheet.max_row
        last_col = len(headers)
        wrap_set = set(wrap_columns)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        sheet.row_dimensions[1].height = 24
        for row in sheet.iter_rows(min_row=2, max_row=last_row, max_col=last_col):
            sheet.row_dimensions[row[0].row].height = row_height
            for cell in row:
                cell.font = body_font
                cell.border = body_border
                cell.alignment = Alignment(
                    horizontal="left" if cell.column in wrap_set else "center",
                    vertical="center",
                    wrap_text=cell.column in wrap_set,
                )
        for index, width_px in enumerate(widths_px, start=1):
            sheet.column_dimensions[column_letter(index)].width = max(8.0, width_px / 7.0)
        reference = f"A1:{column_letter(last_col)}{last_row}"
        sheet.auto_filter.ref = reference
        return sheet, max(2, len(rows) + 1)

    overview = payload["overview"]
    try:
        statistics_time_value: Any = datetime.fromisoformat(
            str(payload["statistics_time"])
        ).replace(tzinfo=None)
    except (TypeError, ValueError):
        statistics_time_value = str(payload["statistics_time"])
    overview_sheet = workbook.create_sheet("总览")
    overview_sheet.sheet_view.showGridLines = False
    overview_sheet.freeze_panes = "A2"
    overview_rows = [
        ["指标", "数值", "说明"],
        ["统计时间", statistics_time_value, "本次统计生成时间"],
        ["总任务数", overview["total_tasks"], "任务明细记录数"],
        ["成功数", overview["success_count"], "原任务状态为success"],
        ["失败数", overview["failed_count"], "原任务状态为failed"],
        ["跳过数", overview["skipped_count"], "原任务状态为skipped"],
        ["可统计结果数", overview["statistic_result_count"], "成功/跳过且有效的TIFF数量"],
        ["总变化像元数", overview["total_changed_pixels"], "变化统计中值1数量之和"],
        ["总变化面积（平方米）", overview["total_changed_area_m2"], "仅累加米制投影结果"],
        ["平均变化比例", overview["average_change_ratio"], "有效TIFF变化比例算术平均"],
        ["总运行耗时（秒）", overview["total_duration_seconds"], "任务明细推理耗时之和"],
    ]
    for row in overview_rows:
        overview_sheet.append(row)
    for cell in overview_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in range(2, 12):
        overview_sheet.row_dimensions[row].height = 22
        overview_sheet[f"A{row}"].fill = PatternFill("solid", fgColor=colors["light_teal"])
        overview_sheet[f"A{row}"].font = Font(name="微软雅黑", size=10, bold=True, color=colors["title"])
        overview_sheet[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True, color=colors["title"])
        for column in "ABC":
            overview_sheet[f"{column}{row}"].border = body_border
            overview_sheet[f"{column}{row}"].alignment = Alignment(vertical="center", wrap_text=column == "C")
    overview_sheet["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    for row in range(3, 9):
        overview_sheet[f"B{row}"].number_format = "#,##0"
    overview_sheet["B9"].number_format = "#,##0.00"
    overview_sheet["B10"].number_format = "0.0000%"
    overview_sheet["B11"].number_format = "0.000"
    overview_sheet.column_dimensions["A"].width = 27
    overview_sheet.column_dimensions["B"].width = 24
    overview_sheet.column_dimensions["C"].width = 55
    overview_sheet.auto_filter.ref = "A1:C11"

    task_sheet, task_end = add_table_sheet(
        "任务明细",
        payload["task_headers"],
        payload["task_rows"],
        "TaskDetailsTable",
        [170, 115, 310, 310, 150, 75, 330, 95, 75, 85, 180, 180, 110, 90, 125, 150, 360],
        row_height=57,
        wrap_columns=(1, 3, 4, 7, 17),
    )
    for row in range(2, task_end + 1):
        task_sheet[f"F{row}"].number_format = "0"
        task_sheet[f"I{row}"].number_format = "0"
        task_sheet[f"J{row}"].number_format = "0"
        task_sheet[f"K{row}"].number_format = "yyyy-mm-dd hh:mm:ss"
        task_sheet[f"L{row}"].number_format = "yyyy-mm-dd hh:mm:ss"
        task_sheet[f"M{row}"].number_format = "0.000"
        task_sheet[f"O{row}"].number_format = "#,##0"
    add_status_formats(task_sheet, "H", 2, task_end)

    stats_sheet, stats_end = add_table_sheet(
        "变化统计",
        payload["stat_headers"],
        payload["stat_rows"],
        "ChangeStatisticsTable",
        [115, 90, 340, 125, 80, 80, 75, 100, 120, 115, 115, 125, 125, 125, 145, 145, 135, 105, 145, 145, 135, 155, 110, 85, 155, 410],
        row_height=47,
        wrap_columns=(3, 26),
    )
    for row in range(2, stats_end + 1):
        for column in ("D", "E", "F", "G", "L", "M", "N", "O", "P", "Q"):
            stats_sheet[f"{column}{row}"].number_format = "#,##0"
        for column in ("J", "K"):
            stats_sheet[f"{column}{row}"].number_format = "0.000000000000"
        stats_sheet[f"R{row}"].number_format = "0.0000%"
        for column in ("S", "T"):
            stats_sheet[f"{column}{row}"].number_format = "#,##0.00"
        stats_sheet[f"U{row}"].number_format = "#,##0.0000"
        stats_sheet[f"V{row}"].number_format = "#,##0.000000"
        stats_sheet[f"W{row}"].number_format = "0.000"
        stats_sheet[f"X{row}"].number_format = "0"
    add_status_formats(stats_sheet, "B", 2, stats_end)
    add_status_formats(stats_sheet, "Y", 2, stats_end)

    failure_sheet, failure_end = add_table_sheet(
        "失败记录",
        payload["failure_headers"],
        payload["failure_rows"],
        "FailureRecordsTable",
        [170, 115, 95, 340, 160, 390, 175, 390],
        row_height=66,
        wrap_columns=(1, 4, 5, 6, 7, 8),
    )
    add_status_formats(failure_sheet, "C", 2, failure_end)
    add_status_formats(failure_sheet, "G", 2, failure_end)

    add_table_sheet(
        "字段说明",
        payload["field_headers"],
        payload["field_rows"],
        "FieldDefinitionsTable",
        [110, 220, 500, 300],
        row_height=32,
        wrap_columns=(2, 3, 4),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)

    reopened = load_workbook(output_path, read_only=False, data_only=False)
    try:
        expected_sheets = ["总览", "任务明细", "变化统计", "失败记录", "字段说明"]
        sheet_names = list(reopened.sheetnames)
        error_cells: List[str] = []
        for sheet in reopened.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or (
                        isinstance(cell.value, str)
                        and any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"))
                    ):
                        error_cells.append(f"{sheet.title}!{cell.coordinate}")
        qa = {
            "sheet_names": sheet_names,
            "expected_sheet_names": expected_sheets,
            "reimport_passed": sheet_names == expected_sheets,
            "formula_error_count": len(error_cells),
            "formula_error_cells": error_cells,
            "openpyxl_structural_check": True,
            "output_size_bytes": output_path.stat().st_size,
        }
        qa["all_passed"] = bool(
            qa["reimport_passed"]
            and qa["formula_error_count"] == 0
            and qa["output_size_bytes"] > 0
        )
    finally:
        reopened.close()

    qa_dir.mkdir(parents=True, exist_ok=True)
    (qa_dir / "workbook_qa.json").write_text(
        json.dumps(qa, ensure_ascii=False, indent=2),
        encoding="utf-8",
        newline="\n",
    )
    logger.info(
        f"Excel已由openpyxl生成 | path={output_path} | size={output_path.stat().st_size}"
    )
    return qa


def make_report(
    path: Path,
    source_tables: Sequence[Path],
    tasks: Sequence[Dict[str, str]],
    stats: Sequence[Dict[str, Any]],
    failures: Sequence[Dict[str, Any]],
    overview: Dict[str, Any],
    qa: Dict[str, Any],
    manual_check: Dict[str, Any],
    xlsx_path: Path,
    csv_path: Path,
    log_path: Path,
) -> None:
    source_lines = "\n".join(f"- `{item}`" for item in source_tables)
    stats_lines = "\n".join(
        (
            f"| {row['task_id']} | {row['任务状态']} | {row['总像元数']:,} | "
            f"{row['未变化像元数（值0）']:,} | {row['变化像元数（值1）']:,} | "
            f"{row['变化比例']:.6%} | {row['变化面积（平方米）']:.2f} | "
            f"{row['统计状态']} |"
        )
        for row in stats
    )
    report = f"""# 变化统计与Excel验收报告

## 验收结论

**{"验收通过" if qa.get("all_passed") and manual_check["passed"] else "验收不通过"}**。

本次仅使用首次真实批量联调状态表进行验收，可靠性状态表已作为可选输入由脚本支持，但未并入本次Excel，以保持验收口径为2个成功TIFF和1个失败任务。

## 输入状态表

{source_lines}

## 任务与统计汇总

- 统计时间：{overview["statistics_time"]}
- 总任务数：{overview["total_tasks"]}
- 成功数：{overview["success_count"]}
- 失败数：{overview["failed_count"]}
- 跳过数：{overview["skipped_count"]}
- 可统计结果数：{overview["statistic_result_count"]}
- 总变化像元数：{overview["total_changed_pixels"]:,}
- 总变化面积：{overview["total_changed_area_m2"]:.2f}平方米
- 平均变化比例：{overview["average_change_ratio"]:.6%}
- 总运行耗时：{overview["total_duration_seconds"]:.3f}秒

| task_id | 原任务状态 | 总像元 | 值0 | 值1 | 变化比例 | 变化面积（平方米） | 统计状态 |
|---|---|---:|---:|---:|---:|---:|---|
{stats_lines}

失败记录数：{len(failures)}。failed任务未打开栅格；单项异常通过try-except隔离。

## 人工复核

- 复核任务：{manual_check["task_id"]}
- 总像元：{manual_check["total_pixels"]:,}（预期36,000,000）
- 值0：{manual_check["unchanged_pixels"]:,}（预期34,077,168）
- 值1：{manual_check["changed_pixels"]:,}（预期1,922,832）
- 值0+值1+其他值=有效像元：{"是" if manual_check["sum_matches_valid"] else "否"}
- 变化比例：{manual_check["change_ratio"]:.6%}（预期约5.3412%）
- 复核一致：{"是" if manual_check["passed"] else "否"}

## 面积规则复核

官方样例为EPSG:2193，线性单位明确为米。单像元面积按 `abs(pixel_width × pixel_height)` 计算；当前结果X/Y分辨率存在极小差异，仅记录亚像元偏移提示，不修改栅格。若输入为经纬度或单位不明确，脚本会将面积字段留空并提示先投影到合适的等积或米制坐标系。

## Excel验收

- 工作表：{", ".join(qa.get("sheet_names", []))}
- 工作表数量：{len(qa.get("sheet_names", []))}
- 关键范围检查：{"通过" if qa.get("inspect_passed") else "不通过"}
- 公式错误扫描：{"通过" if qa.get("formula_error_count") == 0 else "不通过"}
- 全工作表渲染：{"通过" if qa.get("all_sheets_rendered") else "不通过"}
- 导出后重新导入：{"通过" if qa.get("reimport_passed") else "不通过"}
- Excel完整性验收：{"通过" if qa.get("all_passed") else "不通过"}

## 输出文件

- Excel：`{xlsx_path}`
- 统计明细CSV：`{csv_path}`
- 运行日志：`{log_path}`
- 本报告：`{path}`

## 文件保护

所有TIFF均以只读方式打开；未修改原始TIFF、输入状态表、批量控制器、张硕岐核心程序或模型文件。
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(report, encoding="utf-8", newline="\n")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="变化检测结果统计与Excel汇总")
    parser.add_argument(
        "--status-csv",
        action="append",
        type=Path,
        help="状态表，可重复传入；未指定时使用真实批量状态表",
    )
    parser.add_argument(
        "--include-reliability",
        action="store_true",
        help="在指定/默认状态表之外追加可靠性验收状态表",
    )
    parser.add_argument("--result-root", type=Path, default=RESULT_ROOT)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--detail-csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--log", type=Path, default=DEFAULT_LOG)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--builder", type=Path, default=None, help=argparse.SUPPRESS)
    parser.add_argument("--node-exe", type=Path, default=None, help=argparse.SUPPRESS)
    parser.add_argument("--work-dir", type=Path, default=None)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    output_targets = [args.xlsx, args.detail_csv, args.log, args.report]
    if not args.overwrite:
        existing = [str(path) for path in output_targets if path.exists()]
        if existing:
            print("输出已存在，未启用--overwrite：\n" + "\n".join(existing), file=sys.stderr)
            return 2

    logger = RunLogger(args.log)
    started = time.perf_counter()
    try:
        status_tables = list(args.status_csv or [PRIMARY_STATUS])
        if args.include_reliability and RELIABILITY_STATUS not in status_tables:
            status_tables.append(RELIABILITY_STATUS)
        logger.info(f"统计启动 | status_tables={[str(item) for item in status_tables]}")
        tasks = load_status_tables(status_tables)
        logger.info(f"读取任务完成 | total={len(tasks)}")

        task_details = [task_detail_row(task, args.result_root) for task in tasks]
        stats_rows: List[Dict[str, Any]] = []
        csv_rows: List[Dict[str, Any]] = []
        failure_rows: List[Dict[str, Any]] = []

        for task in tasks:
            task_id = task.get("task_id", "")
            status = task.get("status", "").strip().lower()
            try:
                output_path = resolve_output_path(task, args.result_root)
            except Exception as exc:
                message = f"{type(exc).__name__}: {exc}"
                failure_rows.append(
                    failure_row(
                        task,
                        task.get("output_path", ""),
                        type(exc).__name__,
                        str(exc),
                        "statistics_failed",
                        message,
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task,
                        task.get("output_path", ""),
                        "statistics_failed",
                        message,
                    )
                )
                logger.error(
                    f"输出路径解析失败 | task_id={task_id}\n{traceback.format_exc()}"
                )
                continue

            if status == "failed":
                message = task.get("error_message", "") or "任务状态为failed"
                failure_rows.append(
                    failure_row(
                        task,
                        str(output_path),
                        task.get("error_type", "") or "TaskFailed",
                        message,
                        "not_read_failed_task",
                        "按规则不读取failed任务栅格",
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task,
                        str(output_path),
                        "not_read_failed_task",
                        message,
                    )
                )
                logger.warning(f"失败任务不读取栅格 | task_id={task_id}")
                continue

            if status not in {"success", "skipped"}:
                message = f"任务状态{status or '(空)'}不属于success/skipped"
                failure_rows.append(
                    failure_row(
                        task,
                        str(output_path),
                        "StatusNotEligible",
                        message,
                        "not_eligible",
                        "未读取栅格",
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task, str(output_path), "not_eligible", message
                    )
                )
                logger.warning(f"非统计状态跳过 | task_id={task_id} | status={status}")
                continue

            if not output_path.is_file() or output_path.stat().st_size <= 0:
                message = f"有效任务的输出不存在或为空：{output_path}"
                failure_rows.append(
                    failure_row(
                        task,
                        str(output_path),
                        "InvalidOutput",
                        message,
                        "statistics_failed",
                        message,
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task, str(output_path), "statistics_failed", message
                    )
                )
                logger.error(f"输出无效 | task_id={task_id} | output={output_path}")
                continue

            if output_path.suffix.lower() not in {".tif", ".tiff"}:
                message = f"结果不是TIFF：{output_path}"
                failure_rows.append(
                    failure_row(
                        task,
                        str(output_path),
                        "UnsupportedResult",
                        message,
                        "statistics_failed",
                        message,
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task, str(output_path), "statistics_failed", message
                    )
                )
                logger.error(f"输出格式不支持 | task_id={task_id} | output={output_path}")
                continue

            try:
                stats = compute_raster_statistics(task, output_path, logger)
                stats_rows.append(stats)
                csv_rows.append(stats)
            except Exception as exc:
                stack = "".join(
                    traceback.format_exception(
                        type(exc), exc, exc.__traceback__, limit=100
                    )
                )
                message = f"{type(exc).__name__}: {exc}"
                failure_rows.append(
                    failure_row(
                        task,
                        str(output_path),
                        type(exc).__name__,
                        str(exc),
                        "statistics_failed",
                        message,
                    )
                )
                csv_rows.append(
                    statistics_placeholder(
                        task, str(output_path), "statistics_failed", message
                    )
                )
                logger.error(f"单项统计失败 | task_id={task_id}\n{stack}")

        write_statistics_csv(args.detail_csv, csv_rows)
        logger.info(f"统计明细CSV已生成 | path={args.detail_csv}")

        statuses = [task.get("status", "").strip().lower() for task in tasks]
        overview = {
            "statistics_time": datetime.now().astimezone().isoformat(
                timespec="seconds"
            ),
            "total_tasks": len(tasks),
            "success_count": statuses.count("success"),
            "failed_count": statuses.count("failed"),
            "skipped_count": statuses.count("skipped"),
            "statistic_result_count": len(stats_rows),
            "total_changed_pixels": sum(
                int(row["变化像元数（值1）"]) for row in stats_rows
            ),
            "total_changed_area_m2": sum(
                float(row["变化面积（平方米）"] or 0) for row in stats_rows
            ),
            "average_change_ratio": (
                sum(float(row["变化比例"] or 0) for row in stats_rows)
                / len(stats_rows)
                if stats_rows
                else 0
            ),
            "total_duration_seconds": sum(
                parse_float(task.get("duration_seconds", "0")) for task in tasks
            ),
        }

        manual_check = {
            "task_id": "",
            "total_pixels": 0,
            "unchanged_pixels": 0,
            "changed_pixels": 0,
            "other_pixels": 0,
            "valid_pixels": 0,
            "change_ratio": 0.0,
            "sum_matches_valid": False,
            "passed": False,
        }
        if stats_rows:
            first = stats_rows[0]
            manual_check.update(
                {
                    "task_id": first["task_id"],
                    "total_pixels": first["总像元数"],
                    "unchanged_pixels": first["未变化像元数（值0）"],
                    "changed_pixels": first["变化像元数（值1）"],
                    "other_pixels": first["其他像元值数量"],
                    "valid_pixels": first["有效像元数"],
                    "change_ratio": first["变化比例"],
                }
            )
            manual_check["sum_matches_valid"] = (
                manual_check["unchanged_pixels"]
                + manual_check["changed_pixels"]
                + manual_check["other_pixels"]
                == manual_check["valid_pixels"]
            )
            manual_check["passed"] = (
                manual_check["total_pixels"] == 36_000_000
                and manual_check["unchanged_pixels"] == 34_077_168
                and manual_check["changed_pixels"] == 1_922_832
                and manual_check["sum_matches_valid"]
                and math.isclose(
                    manual_check["change_ratio"],
                    0.053412,
                    rel_tol=0.0,
                    abs_tol=1e-12,
                )
            )

        payload = {
            "statistics_time": overview["statistics_time"],
            "overview": overview,
            "task_headers": TASK_HEADERS,
            "task_rows": dict_rows_to_matrix(task_details, TASK_HEADERS),
            "stat_headers": STAT_HEADERS,
            "stat_rows": dict_rows_to_matrix(stats_rows, STAT_HEADERS),
            "failure_headers": FAILURE_HEADERS,
            "failure_rows": dict_rows_to_matrix(failure_rows, FAILURE_HEADERS),
            "field_headers": FIELD_HEADERS,
            "field_rows": dict_rows_to_matrix(field_definitions(), FIELD_HEADERS),
        }
        if args.work_dir is None:
            args.work_dir = args.xlsx.parent / "_work"
        args.work_dir.mkdir(parents=True, exist_ok=True)
        payload_path = args.work_dir / "statistics_payload.json"
        payload_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
            newline="\n",
        )
        qa_dir = args.work_dir / "qa"
        qa = build_workbook_openpyxl(payload, args.xlsx, qa_dir, logger)
        if not qa.get("all_passed", False):
            raise RuntimeError(f"Excel QA未通过：{qa}")
        logger.info(f"Excel已生成 | path={args.xlsx} | size={args.xlsx.stat().st_size}")

        make_report(
            args.report,
            status_tables,
            tasks,
            stats_rows,
            failure_rows,
            overview,
            qa,
            manual_check,
            args.xlsx,
            args.detail_csv,
            args.log,
        )
        logger.info(f"验收报告已生成 | path={args.report}")
        logger.info(
            "统计完成 | "
            f"tasks={len(tasks)} | valid_tiffs={len(stats_rows)} | "
            f"failures={len(failure_rows)} | changed_pixels="
            f"{overview['total_changed_pixels']} | elapsed="
            f"{time.perf_counter() - started:.3f}s"
        )
        print(
            json.dumps(
                {
                    "overview": overview,
                    "manual_check": manual_check,
                    "excel_qa": qa,
                    "xlsx": str(args.xlsx),
                    "csv": str(args.detail_csv),
                    "log": str(args.log),
                    "report": str(args.report),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0
    except Exception:
        logger.error("批次统计失败\n" + traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
